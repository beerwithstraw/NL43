"""
Excel Writer for NL-43 (Rural & Social Obligations).

Master_Data: one row per LOB per Segment (Rural/Social).
Verification sheet: matches PDF layout — rows=LOBs, cols=metrics,
  with Rural and Social as adjacent column pairs.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import (
    MASTER_COLUMNS,
    EXTRACTOR_VERSION,
    NUMBER_FORMAT,
    LOW_CONFIDENCE_FILL_COLOR,
    company_key_to_pascal,
)
from config.row_registry import ROW_ORDER, ROW_DISPLAY_NAMES
from config.lob_registry import LOB_ORDER, LOB_DISPLAY_NAMES
from config.company_metadata import get_metadata
from config.lob_metadata import get_lob_particulars, get_grouped_lob
from extractor.models import CompanyExtract

logger = logging.getLogger(__name__)

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_META_FILL    = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_RURAL_FILL   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_SOCIAL_FILL  = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
_YELLOW_FILL  = PatternFill(
    start_color=LOW_CONFIDENCE_FILL_COLOR,
    end_color=LOW_CONFIDENCE_FILL_COLOR,
    fill_type="solid",
)

_METRIC_COLUMNS = {c for c in MASTER_COLUMNS if c.lower() in ROW_ORDER}


def _year_code_to_fy_end(year_code: str) -> str:
    s = str(year_code).strip()
    if len(s) == 8:
        return s[4:]
    if len(s) == 6:
        return f"20{s[4:]}"
    return s


# ---------------------------------------------------------------------------
# Master_Data sheet
# ---------------------------------------------------------------------------

def _write_master_data(ws, extractions: List[CompanyExtract],
                       existing_rows: Optional[List[list]] = None):
    for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN
    ws.freeze_panes = "A2"

    current_row = 2

    if existing_rows:
        for row_data in existing_rows:
            for col_idx, val in enumerate(row_data, 1):
                if col_idx > len(MASTER_COLUMNS):
                    break
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                if MASTER_COLUMNS[col_idx - 1] in _METRIC_COLUMNS:
                    cell.number_format = NUMBER_FORMAT
            current_row += 1

    for extract in extractions:
        meta = get_metadata(extract.company_key)
        period_data = extract.current_year
        if not period_data:
            continue

        for lob in LOB_ORDER:
            if lob not in period_data.data:
                continue
            lob_data = period_data.data[lob]

            for segment in ("Rural", "Social"):
                seg_key = segment.lower()
                metadata = {
                    "LOB_PARTICULARS":      get_lob_particulars(lob),
                    "Grouped_LOB":          get_grouped_lob(lob),
                    "Company_Name":         meta["company_name"],
                    "Company":              meta["sorted_company"],
                    "NL":                   extract.form_type,
                    "Quarter":              extract.quarter,
                    "Year":                 _year_code_to_fy_end(extract.year),
                    "Segment":              segment,
                    "Sector":               meta["sector"],
                    "Industry_Competitors": meta["competitors"],
                    "GI_Companies":         "GI Company",
                    "Source_File":          extract.source_file,
                }

                row_values = []
                for col_name in MASTER_COLUMNS:
                    if col_name in metadata:
                        row_values.append(metadata[col_name])
                    elif col_name.lower() in ROW_ORDER:
                        val = lob_data.get(col_name.lower(), {}).get(seg_key)
                        row_values.append(val)
                    else:
                        row_values.append(None)

                for col_idx, val in enumerate(row_values, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    if MASTER_COLUMNS[col_idx - 1] in _METRIC_COLUMNS:
                        cell.number_format = NUMBER_FORMAT
                current_row += 1


# ---------------------------------------------------------------------------
# Verification sheet — matches PDF layout
# Rows = LOBs, Cols = Metrics, Rural/Social as sub-column pairs
# ---------------------------------------------------------------------------

def _write_verification_sheet(ws, extract: CompanyExtract):
    ws.cell(row=1, column=1, value=f"VERIFICATION SHEET: {extract.company_name}") \
      .font = Font(bold=True, size=14)
    ws.cell(row=2, column=1,
            value=f"Quarter: {extract.quarter} | Year: {extract.year} | Source: {extract.source_file}")

    if not extract.current_year:
        ws.cell(row=4, column=1, value="No data extracted.").font = Font(italic=True)
        return

    active_lobs = [lob for lob in LOB_ORDER if lob in extract.current_year.data]
    _write_pdf_table(ws, extract.current_year, active_lobs, start_row=4)


def _write_pdf_table(ws, period_data, active_lobs: list, start_row: int):
    """
    PDF-matching layout:
      Col A: Sl.No.         (merged across Rural+Social rows per LOB)
      Col B: Line of Business (merged across Rural+Social rows per LOB)
      Col C: Particular     ("Rural" / "Social")
      Col D: No. of Policies Issued
      Col E: Premium Collected
      Col F: Sum Assured

    Each LOB occupies two consecutive rows: Rural then Social.
    Sl.No. and LOB name cells are merged vertically across the pair.
    """
    # Title row
    title_cell = ws.cell(row=start_row, column=1,
                         value="Rural & Social Obligations (Quarterly Returns)")
    title_cell.font = Font(bold=True, color="FFFFFF", size=11)
    title_cell.fill = _HEADER_FILL
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=6)
    title_cell.alignment = _CENTER_ALIGN

    # Header row
    h = start_row + 1
    headers = [
        "Sl.No.",
        "Line of Business",
        "Particular",
    ] + [ROW_DISPLAY_NAMES.get(m, m) for m in ROW_ORDER]

    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=h, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN

    # Column widths
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    # Data rows — two rows per LOB (Rural then Social)
    data_start = h + 1
    for lob_idx, lob in enumerate(active_lobs):
        rural_row  = data_start + lob_idx * 2
        social_row = rural_row + 1

        lob_data  = period_data.data.get(lob, {})
        lob_label = LOB_DISPLAY_NAMES.get(lob, lob)

        # Sl.No. (1-based, merged)
        slno_cell = ws.cell(row=rural_row, column=1, value=lob_idx + 1)
        slno_cell.alignment = _CENTER_ALIGN
        slno_cell.fill = _META_FILL
        ws.merge_cells(start_row=rural_row, start_column=1,
                       end_row=social_row,  end_column=1)

        # LOB name (merged)
        lob_cell = ws.cell(row=rural_row, column=2, value=lob_label)
        lob_cell.font = Font(bold=True)
        lob_cell.fill = _META_FILL
        lob_cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=rural_row, start_column=2,
                       end_row=social_row,  end_column=2)
        ws.row_dimensions[rural_row].height  = 14
        ws.row_dimensions[social_row].height = 14

        # Particular + metric values for Rural and Social
        for seg_label, row_num, fill in [
            ("Rural",  rural_row,  _RURAL_FILL),
            ("Social", social_row, _SOCIAL_FILL),
        ]:
            seg_key = seg_label.lower()

            part_cell = ws.cell(row=row_num, column=3, value=seg_label)
            part_cell.fill = fill
            part_cell.font = Font(bold=True)
            part_cell.alignment = _CENTER_ALIGN

            for m_idx, metric_key in enumerate(ROW_ORDER):
                val = lob_data.get(metric_key, {}).get(seg_key)
                val_cell = ws.cell(row=row_num, column=4 + m_idx, value=val)
                val_cell.number_format = NUMBER_FORMAT
                val_cell.fill = fill


# ---------------------------------------------------------------------------
# Meta sheet
# ---------------------------------------------------------------------------

def _write_meta_sheet(ws, extractions: List[CompanyExtract], stats: Dict[str, Any]):
    companies = sorted({e.company_name for e in extractions})
    quarters  = sorted({f"{e.quarter}_{e.year}" for e in extractions})

    data = [
        ["Key", "Value"],
        ["extraction_date",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["extractor_version",  EXTRACTOR_VERSION],
        ["files_processed",    stats.get("files_processed", 0)],
        ["files_succeeded",    stats.get("files_succeeded", 0)],
        ["files_failed",       stats.get("files_failed", 0)],
        ["companies",          ", ".join(companies)],
        ["quarters",           ", ".join(quarters)],
    ]
    for r_idx, row in enumerate(data, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
            else:
                cell.fill = _META_FILL


def _sheet_name_for(extract: CompanyExtract) -> str:
    name = f"{company_key_to_pascal(extract.company_key)}_{extract.quarter}_{extract.year}"
    return name[:31]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def save_workbook(extractions: List[CompanyExtract], output_path: str,
                  stats: Optional[Dict[str, Any]] = None):
    if stats is None:
        stats = {}

    output_file = Path(output_path)
    existing_rows = []

    if output_file.exists():
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(output_path)
        new_files = {e.source_file for e in extractions}

        if "Master_Data" in wb.sheetnames:
            ws_old = wb["Master_Data"]
            headers = [cell.value for cell in ws_old[1]]
            if headers[:len(MASTER_COLUMNS)] == MASTER_COLUMNS:
                try:
                    sf_idx = headers.index("Source_File")
                except ValueError:
                    sf_idx = None
                if sf_idx is not None:
                    for row in ws_old.iter_rows(min_row=2, values_only=True):
                        if row[sf_idx] and row[sf_idx] not in new_files:
                            existing_rows.append(list(row))
            else:
                logger.warning("Existing Master_Data has different column layout — regenerating.")
            del wb["Master_Data"]

        for extract in extractions:
            sn = _sheet_name_for(extract)
            if sn in wb.sheetnames:
                del wb[sn]
        if "_meta" in wb.sheetnames:
            del wb["_meta"]
    else:
        wb = Workbook()
        wb.remove(wb.active)

    ws_master = wb.create_sheet("Master_Data", 0)
    _write_master_data(ws_master, extractions, existing_rows=existing_rows)

    for extract in extractions:
        ws = wb.create_sheet(title=_sheet_name_for(extract))
        _write_verification_sheet(ws, extract)

    ws_meta = wb.create_sheet(title="_meta")
    _write_meta_sheet(ws_meta, extractions, stats)

    wb.save(output_path)
    logger.info(f"Excel workbook saved to {output_path}")


def write_validation_summary_sheet(report_path: str, master_path: str, force_company: str = None):
    import pandas as pd
    df = pd.read_csv(report_path)
    summary = df.pivot_table(
        index=["company", "quarter", "year"],
        columns="status", aggfunc="size", fill_value=0
    ).reset_index()
    for col in ["PASS", "WARN", "FAIL", "SKIP"]:
        if col not in summary.columns:
            summary[col] = 0
    summary["Total_Checks"] = summary[["PASS", "SKIP", "WARN", "FAIL"]].sum(axis=1)
    summary = summary.rename(columns={"company": "Company", "quarter": "Quarter", "year": "Year"})
    cols = ["Company", "Quarter", "Year", "Total_Checks", "PASS", "SKIP", "WARN", "FAIL"]
    summary = summary[cols]
    if force_company:
        try:
            existing = pd.read_excel(master_path, sheet_name="Validation_Summary")
            companies_in_new = set(summary["Company"].unique())
            existing = existing[~existing["Company"].isin(companies_in_new)]
            summary = pd.concat([existing, summary], ignore_index=True)
        except Exception:
            pass
    with pd.ExcelWriter(master_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as w:
        summary.to_excel(w, sheet_name="Validation_Summary", index=False)


def write_validation_detail_sheet(report_path: str, master_path: str, force_company: str = None):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    df = pd.read_csv(report_path)
    cols_map = {
        "company": "Company", "quarter": "Quarter", "year": "Year",
        "lob": "LOB", "period": "Period", "check_name": "Check_Name",
        "status": "Status", "expected": "Expected", "actual": "Actual",
        "delta": "Delta", "note": "Note",
    }
    detail = df[df["status"].isin(["FAIL", "WARN"])].copy()
    if detail.empty:
        detail = pd.DataFrame(columns=list(cols_map.values()))
    else:
        detail = detail.rename(columns=cols_map)[list(cols_map.values())]
        detail = detail.sort_values("Status").reset_index(drop=True)

    if force_company:
        try:
            run_companies = set(pd.read_csv(report_path)["company"].unique())
            existing_detail = pd.read_excel(master_path, sheet_name="Validation_Detail")
            if "Company" in existing_detail.columns:
                existing_detail = existing_detail[~existing_detail["Company"].isin(run_companies)]
            detail = pd.concat([existing_detail, detail], ignore_index=True)
        except Exception:
            pass
    with pd.ExcelWriter(master_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as w:
        detail.to_excel(w, sheet_name="Validation_Detail", index=False)

    wb = load_workbook(master_path)
    ws = wb["Validation_Detail"]
    red_fill    = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    status_col = list(cols_map.values()).index("Status") + 1
    for row_idx in range(2, ws.max_row + 1):
        val  = ws.cell(row=row_idx, column=status_col).value
        fill = red_fill if val == "FAIL" else yellow_fill
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
    wb.save(master_path)
    logger.info(f"Validation_Detail sheet written to {master_path}")
